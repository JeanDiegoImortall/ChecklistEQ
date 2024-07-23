import os
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.spinner import Spinner
from kivy.uix.scrollview import ScrollView
from kivy.uix.checkbox import CheckBox
from kivy.uix.textinput import TextInput
from kivy.uix.gridlayout import GridLayout
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from datetime import datetime
import win32com.client as win32
from kivy.core.window import Window
from kivy.uix.label import Label
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from kivy.graphics import Color, Rectangle
from kivy.uix.widget import Widget

Window.clearcolor=(14/255,61/255,76/255,1)
Window.size =(400,600)

class Relatorios(App):
    def build(self):
        self.nconforme_textinputs = []  # Lista para armazenar os TextInputs "Não Conforme"
        self.layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        self.carregar_layout_inicial()

        return self.layout
        
    def carregar_layout_inicial(self):
         # Adicionar botão de reinício no canto superior direito
        self.botao_reiniciar = Button(text='Reiniciar', size_hint=(None, None), size=(100, 30),
                                      pos_hint={'right': 1, 'top': 1})
        self.botao_reiniciar.bind(on_press=self.reiniciar_sistema)
        self.layout.add_widget(self.botao_reiniciar)
        
        # Especifique o nome da planilha que deseja carregar
        sheet_name = 'ID'  # Nome da planilha desejada
        sheet_name2 = 'Operadores'  # Nome da planilha desejada

        # Carrega o arquivo Excel
        self.wb = load_workbook('BD.xlsx')
    
        # Verifica se a planilha especificada existe no arquivo Excel
        if sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]

            # Lê os dados da primeira linha como strings, ignorando a primeira coluna
            first_row = []
            for row in sheet.iter_rows(min_row=1, max_row=1, min_col=2, max_col=sheet.max_column, values_only=True):
                first_row.extend([str(cell) for cell in row])
        
        # Verifica se a planilha especificada existe no arquivo Excel
        if sheet_name2 in self.wb.sheetnames:
            sheet = self.wb[sheet_name2]
            # Lê os dados da primeira linha como strings, ignorando a primeira coluna
            first_row2 = []
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
                first_row2.extend([str(cell) for cell in row])

            # Adiciona o rótulo da planilha (no topo)
            self.layout.add_widget(Label(text=f'Relatório de Inspeção', size_hint=(1, None), height='30sp'))

            # Cria a primeira lista suspensa operador
            self.item_spinner3 = Spinner(
                text='Operador',
                values=first_row2,
                size_hint=(None, None),
                size=(200, 44),
                pos_hint={'center_x': 0.5}
            )

            # Adiciona a primeira lista suspensa ao layout
            self.layout.add_widget(self.item_spinner3)

            # Adiciona o evento de seleção da lista suspensa de operadores
            self.item_spinner3.bind(text=self.atualizar_usuario)
            
            # Cria o TextInput para "Usuário"
            self.usuario_input = TextInput(multiline=False, hint_text='Selecione o nome do inspetor', readonly=True)
            usuario_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
            usuario_layout.add_widget(Label(text='Inspetor:', size_hint_x=None, width='100sp'))
            usuario_layout.add_widget(self.usuario_input)
            self.layout.add_widget(usuario_layout)

            # Cria o TextInput para "nivel"
            self.nivel = TextInput(multiline=False, hint_text='Tipo', readonly=True)
            nivel_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
            nivel_layout.add_widget(Label(text='Inspeção:', size_hint_x=None, width='100sp'))
            nivel_layout.add_widget(self.nivel)
            self.layout.add_widget(nivel_layout)

            # Cria o TextInput para "Senha"
            self.senha_label = Label(text='Senha:', size_hint_x=None, width='100sp')
            self.senha_input = TextInput(password=True, multiline=False, hint_text='Digite a senha')
            senha_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
            senha_layout.add_widget(self.senha_label)
            senha_layout.add_widget(self.senha_input)
            self.layout.add_widget(senha_layout)

            # Armazena o layout de senha para remoção posterior
            self.senha_layout = senha_layout

   

        #return self.layout

    def reiniciar_sistema(self, instance):
        
        self.layout.clear_widgets()  # Limpa todos os widgets atuais
        self.carregar_layout_inicial()  # Reconstrói a interface
                
    def atualizar_usuario(self, spinner, text):
        # Atualiza o TextInput do usuário com o valor selecionado no spinner
        self.usuario_input.text = text
        
        self.layout.remove_widget(self.item_spinner3)

        # Carrega a planilha sheet_name2
        sheet_name2 = 'Operadores'
        if sheet_name2 in self.wb.sheetnames:
            sheet = self.wb[sheet_name2]
            
            # Procura pelo valor selecionado (text) na coluna 1
            row_index = None
            for row in range(2, sheet.max_row + 1):  # Começa de 2 para ignorar o cabeçalho
                if sheet.cell(row=row, column=1).value == text:
                    row_index = row
                    break
            
            # Se encontrou o valor, atualiza o TextInput nivel com o valor da coluna 2
            if row_index is not None:
                nivel_value = sheet.cell(row=row_index, column=2).value
                self.tipo=int(nivel_value)
                nivel_value2 = sheet.cell(row=row_index, column=4).value
                self.nivel.text = str(nivel_value2)
                
                
         # Adiciona o botão de Entrar
        self.botao_entrar = Button(text='Entrar', size_hint=(None, None), size=(150, 50), pos_hint={'center_x': 0.5})
        self.botao_entrar.bind(on_press=self.on_botao_entrar_press)
        self.layout.add_widget(self.botao_entrar)

        return self.layout

    def on_botao_entrar_press(self, instance):
        usuario = self.usuario_input.text
        senha = self.senha_input.text
        self.entrar(usuario, senha)

    def entrar(self, text, senha):
        # Atualiza o TextInput da senha com o valor selecionado no spinner
        self.usuario_input.text = text
        self.senha_input.text = senha

        # Carrega a planilha sheet_name2
        sheet_name2 = 'Operadores'
        if sheet_name2 in self.wb.sheetnames:
            sheet = self.wb[sheet_name2]

            # Procura pelo valor selecionado (text) na coluna 1
            row_index = None
            for row in range(2, sheet.max_row + 1):  # Começa de 2 para ignorar o cabeçalho
                if sheet.cell(row=row, column=1).value == text:
                    row_index = row
                    break

            # Se encontrou o valor, verifica se a senha é correta
            if row_index is not None:
                senha_correta = str(sheet.cell(row=row_index, column=3).value)
                if senha != senha_correta:
                    self.mostrar_alerta("Senha incorreta. Tente novamente.")
                    return  # Para o código se a senha estiver incorreta
                else:
                    #self.nivel_input.text = str(sheet.cell(row=row_index, column=2).value)
                    self.layout.remove_widget(self.item_spinner3)
                    self.layout.remove_widget(self.senha_layout)
                    self.layout.remove_widget(self.botao_entrar) 
                    self.novo_layout()  

    def mostrar_alerta(self, mensagem):
        conteudo = BoxLayout(orientation='vertical')
        conteudo.add_widget(Label(text=mensagem))
        botao_fechar = Button(text='Fechar', size_hint=(None, None), size=(100, 50), pos_hint={'center_x': 0.5})
        conteudo.add_widget(botao_fechar)

        popup = Popup(title='Alerta', content=conteudo, size_hint=(None, None), size=(400, 200))
        botao_fechar.bind(on_release=popup.dismiss)
        popup.open()

    def novo_layout(self):
        # Verifica o valor do TextInput nivel
        if self.tipo == 1:
            # Cria um novo TextInput para chassi
            self.chassi_input = TextInput(multiline=False, hint_text='Digite o número do chassi')
            chassi_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
            chassi_layout.add_widget(Label(text='Chassi:', size_hint_x=None, width='100sp'))
            chassi_layout.add_widget(self.chassi_input)
            self.layout.add_widget(chassi_layout)

            # Adiciona o botão de inspecionar
            self.botao_chassi = Button(text='Inspecionar', size_hint=(None, None), size=(150, 50), pos_hint={'center_x': 0.5})
            self.botao_chassi.bind(on_press=self.inspecionar)
            self.layout.add_widget(self.botao_chassi)

            sheet_name2 = 'Chassis'
            if sheet_name2 in self.wb.sheetnames:
                sheet = self.wb[sheet_name2]

            # Procura pelo valor selecionado (text) na coluna 1
                index = None
                index =sheet.max_row + 1
                today=datetime.today().strftime('%d%m%y')
                serial_number = f'E{today}{index}'    

            self.chassi_input.text=serial_number
            # Cria um novo TextInput para fornecedor
            #self.fornecedor_input = TextInput(multiline=False, hint_text='Digite nome do fornecedor')
            #fornecedor_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
            #fornecedor_layout.add_widget(Label(text='Fornecedor:', size_hint_x=None, width='100sp'))
            #fornecedor_layout.add_widget(self.fornecedor_input)
            #self.layout.add_widget(fornecedor_layout)

        if self.tipo == 2:
            self.novo_layout2()
        if self.tipo >= 3:
            self.novo_layout3()
    
    def inspecionar(self, instance):
        
        selected_planilha2 = 'Fatiadora'
        selected_coluna2 = 3

        if selected_planilha2 in self.wb.sheetnames:
            sheet = self.wb[selected_planilha2]

            # Encontra a coluna correspondente ao valor selecionado na segunda lista
            col_index = 3
            
            # Se encontrou a coluna, cria os componentes (checklist ou textbox)
            if col_index is not None:
                self.adicionar_componentes2(sheet, col_index, selected_planilha2, selected_coluna2)

    def update_border(self, instance, value):
        instance.canvas.before.clear()
        with instance.canvas.before:
            Color(0, 0, 0.1, 1)  
            Rectangle(size=instance.size, pos=instance.pos)
    
    def adicionar_componentes2(self, sheet, col_index, selected_planilha2, selected_coluna2):
        components_layout = GridLayout(cols=1, size_hint_y=None)
        components_layout.bind(minimum_height=components_layout.setter('height'))

        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=col_index).value
            cell_value2 = sheet.cell(row=row, column=1).value
            if sheet.cell(row=row, column=1).value != None:
                cell_value2 = int(cell_value2)   
                        
            if cell_value == 'x' and cell_value2 == int(self.tipo):
                # Cria um checklist
                item_layout = BoxLayout(orientation='vertical', size_hint_y=None, height='100sp')
                # Adicionando borda ao layout
                with item_layout.canvas.before:
                    Color(1, 1, 1, 1)  
                    self.border = Rectangle(size=item_layout.size, pos=item_layout.pos)

                # Atualizando a posição e o tamanho da borda quando o layout é redimensionado
                item_layout.bind(pos=self.update_border, size=self.update_border)

                item_text = Label(
                    text=str(sheet.cell(row=row, column=2).value), 
                    halign='left', 
                    valign='center', 
                    size_hint_y=2.3,
                    size_hint_x=1, 
                    text_size=(self.layout.width * 0.9, None),
                    padding=(10, 0)
                    
                )
                group_name = f"group_{row}"
                self.item_checkbox = CheckBox(group=group_name, size_hint_x=1,width='50sp')
                self.item_checkbox2 = CheckBox(group=group_name, size_hint_x=1,width='50sp')
                item_layout.add_widget(item_text)
                item_layout.add_widget(self.item_checkbox)
                item_layout.add_widget(Widget(size_hint_x=None, width=0))
                item_layout.add_widget(self.item_checkbox2)
                components_layout.add_widget(item_layout)

                 # Define o evento de adicionar/remover a label 'OK' ao lado do checkbox
                def on_checkbox_active(checkbox, value, layout=item_layout):
                    if value:
                        layout.add_widget(Label(text='OK', size_hint_x=1))
                    else:
                        for widget in layout.children:
                            if isinstance(widget, Label) and widget.text == 'OK':
                                layout.remove_widget(widget)
                                break

                self.item_checkbox.bind(active=on_checkbox_active)

                
                # Define o evento de adicionar/remover a label 'OK' ao lado do checkbox
                def on_checkbox_active2(checkbox, value, layout=item_layout):
                    if value:
                        if value:
                            layout.add_widget(Label(text='Não Conforme', size_hint_x=1,color=(1, 1, 0.2, 1)))
                    else:
                        for widget in layout.children:
                            if isinstance(widget, Label) and widget.text == 'Não Conforme':
                                layout.remove_widget(widget)
                                break

                self.item_checkbox2.bind(active=on_checkbox_active2)

            elif cell_value == 'dx' and cell_value2 == int(self.tipo):
                # Cria um textbox
                item_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
                item_label = Label(text=str(sheet.cell(row=row, column=2).value), halign='right', valign='middle', size_hint_x=None, width='300sp')
                item_textbox = TextInput(multiline=False, size_hint_x=None, width='100sp')
                item_layout.add_widget(item_label)
                item_layout.add_widget(item_textbox)
                components_layout.add_widget(item_layout)
                                
        scroll_view = ScrollView(size_hint=(1, 1))
        scroll_view.add_widget(components_layout)
        self.layout.add_widget(scroll_view)
        self.layout.remove_widget(self.botao_chassi)

         # Cria um novo TextInput para observação
        self.observacao_input = TextInput(multiline=False, hint_text='Observações')
        observacao_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
        observacao_layout.add_widget(Label(text='Obs:', size_hint_x=None, width='100sp'))
        observacao_layout.add_widget(self.observacao_input)
        self.layout.add_widget(observacao_layout)

         # Adiciona o botão de exportar PDF após o checklist ser gerado
        export_button = Button(text='Exportar Relatório', size_hint=(None, None), size=(150, 50), pos_hint={'center_x': 0.5})
        export_button.bind(on_press=lambda instance: self.exportar_e_enviar_relatorio2(selected_planilha2, selected_coluna2))
        self.layout.add_widget(export_button)
    
    def exportar_e_enviar_relatorio2(self, planilha, coluna):
        if not self.verificar_campos_preenchidos2():
            # Mostrar mensagem de erro
            popup = Popup(title='Campos não preenchidos',
                          content=Label(text='Por favor, preencha todos os campos antes de exportar.'),
                          size_hint=(None, None), size=(400, 200))
            popup.open()
            return
        
        # Cria o nome do arquivo PDF com base no tempo
        now = datetime.now()
        dt_string = now.strftime("%Y%m%d_%H%M%S")
        file_name = f'{self.chassi_input.text}-Relatorio.pdf'
        pdf_path = os.path.join(os.getcwd(), file_name)

        # Verifica se o arquivo PDF já existe
        if os.path.exists(pdf_path):
            # Mostrar mensagem de alerta
            popup = Popup(title='Relatório já Existente',
                          content=Label(text=f'O relatório {file_name} já existe.'),
                          size_hint=(None, None), size=(400, 200))
            popup.open()
            return

        # Gera o PDF
        c = canvas.Canvas(pdf_path)
        self.gerar_pdf2(c)

        # Salva o PDF
        c.save()

        # Envia o e-mail com o PDF anexado
        self.enviar_email_outlook2(pdf_path)

        # Exibir mensagem de arquivo criado
        popup = Popup(title='Relatório Criado',
                      content=Label(text=f'Relatório salvo como {file_name} e enviado por e-mail.'),
                      size_hint=(None, None), size=(400, 200))
        popup.open()

       
        # Adicionar dados no Excel
        workbook_path = 'BD.xlsx'  
        workbook = load_workbook(workbook_path)
        sheet = workbook['Chassis']

        # Encontrar a próxima linha vazia
        next_row = sheet.max_row + 1

        # Adicionar dados na próxima linha vazia
        sheet.cell(row=next_row, column=1, value=str(self.chassi_input.text))
        sheet.cell(row=next_row, column=6, value=int("1"))
        # Salvar o workbook
        workbook.save(workbook_path)
        workbook.close()

        self.layout.clear_widgets()  # Limpa todos os widgets atuais
        self.carregar_layout_inicial()  # Reconstrói a interface

    def verificar_campos_preenchidos2(self):
        if not self.usuario_input.text:
            return False

        for widget in self.layout.children:
            if isinstance(widget, ScrollView):
                components_layout = widget.children[0]
                for item_layout in components_layout.children:
                    if isinstance(item_layout, BoxLayout):
                        checkboxes = []
                        for child in item_layout.children:
                            # Adiciona CheckBox ao grupo
                            if isinstance(child, CheckBox):
                                    checkboxes.append(child)

                            # Verifica o par de CheckBox
                            if len(checkboxes) == 2:
                               if not checkboxes[0].active and not checkboxes[1].active:
                                  return False    

                            if isinstance(child, TextInput) and not child.text:
                                return False
                 # Verifica todos os TextInputs "Não Conforme"
                for nconforme_textinput in self.nconforme_textinputs:
                    if not nconforme_textinput.text:
                         return False

        return True
    
    def gerar_pdf2(self, c):
        # Adicionar cabeçalho
        header_text = 'Inspeção de Chassis'
        c.setFont("Helvetica-Bold", 16)
        c.drawString(100, 800, f'{header_text}')

        # Linha com texto "Relatório de Inspeção"
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredString(300, 740, "Relatório de Inspeção Chassi")

        # Adicionar uma linha
        c.line(30, 730, 550, 730)

        # Usuário do chassi no PDF
        c.setFont("Helvetica", 12)
        c.drawString(100, 780, f'Inspetor: {self.usuario_input.text}')
        c.drawString(100, 760, f'Número do Chassis: {self.chassi_input.text}')

        # Obter todas as labels dos checkboxes marcados e labels dos textboxes com valores
        self.nok=None
        labels_checked = []
        for widget in self.layout.children:
            if isinstance(widget, ScrollView):
                components_layout = widget.children[0]
                for item_layout in components_layout.children:
                    if isinstance(item_layout, BoxLayout):
                        for child in item_layout.children:
                            if isinstance(child, CheckBox) and child.active:
                                # Adicionar a label associada ao checkbox marcado
                                for sub_child in item_layout.children:
                                    if isinstance(sub_child, Label):
                                        labels_checked.append(sub_child.text)
                            elif isinstance(child, TextInput):
                                # Adicionar a label e o valor do textbox
                                for sub_child in item_layout.children:
                                    if isinstance(sub_child, Label):
                                        labels_checked.append(f"{sub_child.text}: {child.text}")

        # Inverter a lista de labels e valores
        if "Não Conforme" in labels_checked:
            self.nok=1
        labels_checked.reverse()

        # Escrever as labels e valores no PDF
        c.setFont("Helvetica", 12)
        y_position = 700
        page_height = 750  # Altura da página para controle de quebra de página
        for label in labels_checked:
            if y_position < 50:  # Se o espaço na página atual acabou, cria uma nova página
                c.showPage()
                c.setFont("Helvetica", 12)
                page_height = 800 # Reinicia a altura da nova página
                y_position = 800
        
            c.drawString(100, y_position, label)
            y_position -= 25  # Espaçamento entre as labels
            page_height -= 20  # Atualiza a altura da página atual
        
        #inserir observacao
        if y_position < 50:  # Se o espaço na página atual acabou, cria uma nova página
                c.showPage()
                c.setFont("Helvetica", 12)
                page_height = 800 # Reinicia a altura da nova página
                y_position = 800
              
        
        c.drawString(100, y_position, f'Obs: {self.observacao_input.text}')

    def enviar_email_outlook2(self, attachment_path):
        try:
            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)
            mail.To = 'processos01@equimatec.ind.br'
            mail.Subject = 'Relatório de Inspeção'
            if self.nok==1:
                mail.Body = 'Segue em anexo o relatório de inspeção gerado. Atenção! Possui não conformidades.'    
            else:
                mail.Body = 'Segue em anexo o relatório de inspeção gerado, sem resalvas'
            mail.Attachments.Add(attachment_path)
            mail.Send()
        except Exception as e:
            # Manipular erros de envio de e-mail aqui
            print(f"Erro ao enviar e-mail: {str(e)}")

    def novo_layout2(self):
        # Especifique o nome da planilha que deseja carregar
        sheet_name = 'Chassis'  # Nome da planilha desejada
        
        # Carrega o arquivo Excel
        self.wb = load_workbook('BD.xlsx')

        # Verifica se a planilha especificada existe no arquivo Excel
        if sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]

            # Lê os dados da primeira linha como strings, ignorando a primeira coluna
            first_row = []
            for row in range(2, sheet.max_row + 1):  # Começa da linha 2 até a última linha da planilha
                    # Obtém o valor da coluna 6 para a linha atual
                    value_col6 = sheet.cell(row=row, column=6).value
                    # Verifica se o valor da coluna 6 é igual a "1" (como string)
                    if value_col6 == 1:
                       first_row.append(str(sheet.cell(row=row, column=1).value))
        
            # Cria a primeira lista suspensa para selecionar um item da primeira linha
            self.item_spinner1 = Spinner(
                text='Escolha um chassi',
                values=first_row,
                size_hint=(None, None),
                size=(200, 44),
                pos_hint={'center_x': 0.5}
            )

            # Adiciona a primeira lista suspensa ao layout
            self.layout.add_widget(self.item_spinner1)

            self.item_spinner1.bind(text=self.atualizar_chassi)
        else:
            # Caso a planilha especificada não seja encontrada
            self.layout.add_widget(Label(text=f'Banco de dados "{sheet_name}" não encontrada.'))   

    def novo_layout3(self):
        # Especifique o nome da planilha que deseja carregar
        sheet_name = 'Chassis'  # Nome da planilha desejada
        
        # Carrega o arquivo Excel
        self.wb = load_workbook('BD.xlsx')

        # Verifica se a planilha especificada existe no arquivo Excel
        if sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]

            # Lê os dados da primeira linha como strings, ignorando a primeira coluna
            first_row4 = []
            for row in range(2, sheet.max_row + 1):  # Começa da linha 2 até a última linha da planilha
                    # Obtém o valor da coluna 6 para a linha atual
                    value_col6 = sheet.cell(row=row, column=6).value
                    value_col6 = value_col6 + 1
                    
                    # Verifica se o valor da coluna 6 é igual a "1" (como string)
                    if value_col6 == int(self.tipo):
                       first_row4.append(str(sheet.cell(row=row, column=3).value))
        
        
            # Cria a primeira lista suspensa para selecionar um item da primeira linha
            self.item_spinner6 = Spinner(
                text='Informe o Nº Série',
                values=first_row4,
                size_hint=(None, None),
                size=(200, 44),
                pos_hint={'center_x': 0.5}
            )

            # Adiciona a primeira lista suspensa ao layout
            self.layout.add_widget(self.item_spinner6)

            self.item_spinner6.bind(text=self.atualizar_inspecao)
        else:
            # Caso a planilha especificada não seja encontrada
            self.layout.add_widget(Label(text=f'Banco de dados "{sheet_name}" não encontrado.')) 

    def atualizar_inspecao(self, spinner, text):

            self.layout.remove_widget(self.item_spinner6)           

         # Cria um novo TextInput para Equipamento
            self.equipamento_input = TextInput(multiline=False, readonly=True)
            equipamento_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
            equipamento_layout.add_widget(Label(text='Equipamento:', size_hint_x=None, width='100sp'))
            equipamento_layout.add_widget(self.equipamento_input)
            self.layout.add_widget(equipamento_layout)
        
        # Cria um novo TextInput para Modelo
            self.modelo_input = TextInput(multiline=False, readonly=True)
            modelo_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
            modelo_layout.add_widget(Label(text='Modelo:', size_hint_x=None, width='100sp'))
            modelo_layout.add_widget(self.modelo_input)
            self.layout.add_widget(modelo_layout)

            # Cria o TextInput para "Número de Série"
            self.numero_serie_input = TextInput(multiline=False, hint_text='Digite o número de série')
            numero_serie_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
            numero_serie_layout.add_widget(Label(text='Nº de Série:', size_hint_x=None, width='100sp'))
            numero_serie_layout.add_widget(self.numero_serie_input)
            self.layout.add_widget(numero_serie_layout)   
        
        # Atualiza o TextInput do equipamento e modelo
        # Carrega a planilha sheet_name2
        # Carrega o arquivo Excel
            valor =text
            self.wb = load_workbook('BD.xlsx')   
            sheet_name2 = 'Chassis'
            if sheet_name2 in self.wb.sheetnames:
                sheet = self.wb[sheet_name2]
                
            # Procura pelo valor selecionado (text) na coluna 3
            row_index = None
            for row in range(1, sheet.max_row + 1):  # Começa de 2 para ignorar o cabeçalho
                 if sheet.cell(row=row, column=3).value == valor:
                    self.equipamento_input.text= str(sheet.cell(row=row, column=4).value)
                    self.modelo_input.text=str(sheet.cell(row=row, column=5).value)
                    self.numero_serie_input.text=str(sheet.cell(row=row, column=3).value)
                    break
            
            self.exibir_componentes3()
           
    def atualizar_chassi(self, spinner, text):
       # Cria o TextInput para "Número do chassi"
        self.numero_chassi_input = TextInput(multiline=False, hint_text='Nº Chassi',readonly=True)
        numero_chassi_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
        numero_chassi_layout.add_widget(Label(text='Nº Chassi:', size_hint_x=None, width='100sp'))
        numero_chassi_layout.add_widget(self.numero_chassi_input)
        self.layout.add_widget(numero_chassi_layout)

        self.numero_chassi_input.text = self.item_spinner1.text

        self.layout.remove_widget(self.item_spinner1)

       #verifica se o relatorio deste chassi esta disponivel
        file_name2 = f'{self.numero_chassi_input.text}-Relatorio.pdf'#nome do relatorio chassi
        pdf_path2 = os.path.join(os.getcwd(), file_name2)

        # Verifica se o arquivo PDF já existe
        if os.path.exists(pdf_path2):
           popup = Popup(title='Relatório',
                          content=Label(text=f'Insira numero de série.'),
                          size_hint=(None, None), size=(400, 200))
           popup.open()
           
            
        else:
            popup = Popup(title='Relatório não encontrado',
                          content=Label(text=f'O relatório do chassi não existe.'),
                          size_hint=(None, None), size=(400, 200))
            popup.open()
            return
       
       
       
        # Cria o TextInput para "Número de Série"
        self.numero_serie_input = TextInput(multiline=False, hint_text='Digite o número de série')
        numero_serie_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
        numero_serie_layout.add_widget(Label(text='Nº de Série:', size_hint_x=None, width='100sp'))
        numero_serie_layout.add_widget(self.numero_serie_input)
        self.layout.add_widget(numero_serie_layout)
                
        # Especifique o nome da planilha que deseja carregar
        sheet_name = 'ID'  # Nome da planilha desejada
        
        # Carrega o arquivo Excel
        self.wb = load_workbook('BD.xlsx')

        # Verifica se a planilha especificada existe no arquivo Excel
        if sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]

            # Lê os dados da primeira linha como strings, ignorando a primeira coluna
            first_row = []
            for row in sheet.iter_rows(min_row=1, max_row=1, min_col=2, max_col=sheet.max_column, values_only=True):
                first_row.extend([str(cell) for cell in row])
        
            # Cria a primeira lista suspensa para selecionar um item da primeira linha
            self.item_spinner4 = Spinner(
                text='Escolha um equipamento',
                values=first_row,
                size_hint=(None, None),
                size=(200, 44),
                pos_hint={'center_x': 0.5}
            )

            # Adiciona a primeira lista suspensa ao layout
            self.layout.add_widget(self.item_spinner4)

            # Cria a segunda lista suspensa (inicialmente vazia)
            self.item_spinner5 = Spinner(
                text='Selecione primeiro',
                size_hint=(None, None),
                size=(200, 44),
                pos_hint={'center_x': 0.5}
            )

            # Adiciona a segunda lista suspensa ao layout
            self.layout.add_widget(self.item_spinner5)

            # Adiciona o evento de seleção da primeira lista suspensa
            self.item_spinner4.bind(text=self.atualizar_lista)

             # Adiciona o botão de inspecionar
            #self.botao_inspecao = Button(text='Inspecionar', size_hint=(None, None), size=(150, 50), pos_hint={'center_x': 0.5})
            #self.botao_inspecao.bind(on_press=self.inspecionar)
            #self.layout.add_widget(self.botao_inspecao)

        else:
            # Caso a planilha especificada não seja encontrada
            self.layout.add_widget(Label(text=f'Planilha "{sheet_name}" não encontrada.'))
            
    def atualizar_lista(self, instance, value):
        # Reseta os valores da segunda lista suspensa
        #self.item_spinner5.unbind(text=self.exibir_componentes)
        self.item_spinner5.text = 'Selecione primeiro'
        self.item_spinner5.values = []

        selected_value = value
        sheet_name = 'ID'  # nome da planilha

        # Carrega o arquivo Excel
        self.wb = load_workbook('BD.xlsx')

        # Verifica se a planilha especificada existe no arquivo Excel
        if sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]

            # Encontra a coluna correspondente ao valor selecionado na primeira lista
            col_index = None
            for col in range(2, sheet.max_column + 1):  # Começa de 2 para ignorar a primeira coluna
                if sheet.cell(row=1, column=col).value == selected_value:
                    col_index = col
                    break

            # Se encontrou a coluna, carrega os valores dela
            if col_index is not None:
                column_values = [str(sheet.cell(row=row, column=col_index).value) for row in range(2, sheet.max_row + 1)]
                column_values = [value for value in column_values if value and value != 'None']  # Remove valores None ou 'None'
                self.item_spinner5.values = column_values
                self.item_spinner5.text = 'Escolha um modelo'

            # Define o evento de seleção da segunda lista suspensa
            self.item_spinner5.bind(text=self.exibir_componentes)

    def exibir_componentes(self, instance, value):
        if value == 'Selecione primeiro' or not value:
            return  # Retorna se a seleção for inválida

        selected_planilha = self.item_spinner4.text
        selected_coluna = value

        
         # Cria um novo TextInput para Equipamento
        self.equipamento_input = TextInput(multiline=False, readonly=True)
        equipamento_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
        equipamento_layout.add_widget(Label(text='Equipamento:', size_hint_x=None, width='100sp'))
        equipamento_layout.add_widget(self.equipamento_input)
        self.layout.add_widget(equipamento_layout)
        
        # Cria um novo TextInput para Modelo
        self.modelo_input = TextInput(multiline=False, readonly=True)
        modelo_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
        modelo_layout.add_widget(Label(text='Modelo:', size_hint_x=None, width='100sp'))
        modelo_layout.add_widget(self.modelo_input)
        self.layout.add_widget(modelo_layout)

        self.equipamento_input.text = self.item_spinner4.text
        self.modelo_input.text=selected_coluna

        self.layout.remove_widget(self.item_spinner4)
        self.layout.remove_widget(self.item_spinner5)


        if selected_planilha in self.wb.sheetnames:
            sheet = self.wb[selected_planilha]

            # Encontra a coluna correspondente ao valor selecionado na segunda lista
            col_index = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == selected_coluna:
                    col_index = col
                    break
        

            # Se encontrou a coluna, cria os componentes (checklist ou textbox)
            if col_index is not None:
                self.adicionar_componentes(sheet, col_index, selected_planilha, selected_coluna)

    def exibir_componentes2(self):
        
        selected_planilha = self.equipamento_input.text
        selected_coluna = self.modelo_input.text
        
        if selected_planilha in self.wb.sheetnames:
            sheet = self.wb[selected_planilha]

            # Encontra a coluna correspondente ao valor selecionado na segunda lista
            col_index = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == selected_coluna:
                    col_index = col
                    break
        

            # Se encontrou a coluna, cria os componentes (checklist ou textbox)
            if col_index is not None:
                self.adicionar_componentes(sheet, col_index, selected_planilha, selected_coluna)

    def exibir_componentes3(self):
        
        selected_planilha = self.equipamento_input.text
        selected_coluna = self.modelo_input.text
        
        if selected_planilha in self.wb.sheetnames:
            sheet = self.wb[selected_planilha]

            # Encontra a coluna correspondente ao valor selecionado na segunda lista
            col_index = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == selected_coluna:
                    col_index = col
                    break
        

            # Se encontrou a coluna, cria os componentes (checklist ou textbox)
            if col_index is not None:
                self.adicionar_componentes3(sheet, col_index, selected_planilha, selected_coluna)

    def adicionar_componentes(self, sheet, col_index, selected_planilha, selected_coluna):
        components_layout = GridLayout(cols=1, size_hint_y=None)
        components_layout.bind(minimum_height=components_layout.setter('height'))

        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=col_index).value
            cell_value2 = sheet.cell(row=row, column=1).value
            if sheet.cell(row=row, column=1).value != None:
                cell_value2 = int(cell_value2)   
                        
            if cell_value == 'x' and cell_value2 == int(self.tipo):
                # Cria um checklist
                item_layout = BoxLayout(orientation='vertical', size_hint_y=None, height='100sp')
                with item_layout.canvas.before:
                    Color(1, 1, 1, 1)  
                    self.border = Rectangle(size=item_layout.size, pos=item_layout.pos)

                # Atualizando a posição e o tamanho da borda quando o layout é redimensionado
                item_layout.bind(pos=self.update_border, size=self.update_border)
                
                item_text = Label(
                    text=str(sheet.cell(row=row, column=2).value), 
                    halign='left', 
                    valign='center', 
                    size_hint_y=2.3,
                    size_hint_x=1.0, 
                     text_size=(self.layout.width * 0.9, None),
                    padding=(10, 0)
                    
                )
                group_name = f"group_{row}"
                self.item_checkbox = CheckBox(group=group_name, size_hint_x=1, width='50sp')
                self.item_checkbox2 = CheckBox(group=group_name, size_hint_x=1, width='50sp')
                item_layout.add_widget(item_text)
                item_layout.add_widget(self.item_checkbox)
                item_layout.add_widget(Widget(size_hint_x=None, width=0))
                item_layout.add_widget(self.item_checkbox2)
                components_layout.add_widget(item_layout)

                 # Define o evento de adicionar/remover a label 'OK' ao lado do checkbox
                def on_checkbox_active(checkbox, value, layout=item_layout):
                    if value:
                        layout.add_widget(Label(text='OK', size_hint_x=1))
                    else:
                        for widget in layout.children:
                            if isinstance(widget, Label) and widget.text == 'OK':
                                layout.remove_widget(widget)
                                break

                self.item_checkbox.bind(active=on_checkbox_active)

                
                # Define o evento de adicionar/remover a label 'OK' ao lado do checkbox
                def on_checkbox_active2(checkbox, value, layout=item_layout):
                    if value:
                        if value:
                            layout.add_widget(Label(text='Não Conforme', size_hint_x=1,color=(1, 1, 0.2, 1)))
                    else:
                        for widget in layout.children:
                            if isinstance(widget, Label) and widget.text == 'Não Conforme':
                                layout.remove_widget(widget)
                                break

                self.item_checkbox2.bind(active=on_checkbox_active2)

            elif cell_value == 'dx' and cell_value2 == int(self.tipo):
                # Cria um textbox
                item_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
                item_label = Label(text=str(sheet.cell(row=row, column=2).value), halign='right', valign='middle', size_hint_x=None, width='300sp')
                item_textbox = TextInput(multiline=False, size_hint_x=None, width='100sp')
                item_layout.add_widget(item_label)
                item_layout.add_widget(item_textbox)
                components_layout.add_widget(item_layout)

        scroll_view = ScrollView(size_hint=(1, 1))
        scroll_view.add_widget(components_layout)
        self.layout.add_widget(scroll_view)

        # Cria um novo TextInput para observação
        self.observacao_input = TextInput(multiline=False, hint_text='Observações')
        observacao_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
        observacao_layout.add_widget(Label(text='Obs:', size_hint_x=None, width='100sp'))
        observacao_layout.add_widget(self.observacao_input)
        self.layout.add_widget(observacao_layout)

        # Adiciona o botão de exportar PDF após o checklist ser gerado
        export_button = Button(text='Exportar Relatório', size_hint=(None, None), size=(150, 50), pos_hint={'center_x': 0.5})
        export_button.bind(on_press=lambda instance: self.exportar_e_enviar_relatorio(selected_planilha, selected_coluna))
        self.layout.add_widget(export_button)

    def adicionar_componentes3(self, sheet, col_index, selected_planilha, selected_coluna):
        components_layout = GridLayout(cols=1, size_hint_y=None)
        components_layout.bind(minimum_height=components_layout.setter('height'))

        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=col_index).value
            cell_value2 = str(sheet.cell(row=row, column=1).value)
            if sheet.cell(row=row, column=1).value != None:
                cell_value2 = int(cell_value2)   
            if cell_value == 'x' and cell_value2 == int(self.tipo):
                # Cria um checklist
                item_layout = BoxLayout(orientation='vertical', size_hint_y=None, height='100sp')
                with item_layout.canvas.before:
                    Color(1, 1, 1, 1)  
                    self.border = Rectangle(size=item_layout.size, pos=item_layout.pos)

                # Atualizando a posição e o tamanho da borda quando o layout é redimensionado
                item_layout.bind(pos=self.update_border, size=self.update_border)

                item_text = Label(
                    text=str(sheet.cell(row=row, column=2).value),
                    halign='left',
                    valign='center',
                    size_hint_y=2.3,
                    size_hint_x=1,
                    text_size=(self.layout.width * 0.9, None),
                    padding=(10, 0)
                )
                group_name = f"group_{row}"
                self.item_checkbox = CheckBox(group=group_name, size_hint_x=1, width='50sp')
                self.item_checkbox2 = CheckBox(group=group_name, size_hint_x=1, width='50sp')
                item_layout.add_widget(item_text)
                item_layout.add_widget(self.item_checkbox)
                item_layout.add_widget(Widget(size_hint_x=None, width=0))
                item_layout.add_widget(self.item_checkbox2)
                components_layout.add_widget(item_layout)

                # Define o evento de adicionar/remover a label 'OK' ao lado do checkbox
                def on_checkbox_active(checkbox, value, layout=item_layout):
                    if value:
                        layout.add_widget(Label(text='OK', size_hint_x=1))
                    else:
                        for widget in layout.children:
                            if isinstance(widget, Label) and widget.text == 'OK':
                                layout.remove_widget(widget)
                                break

                self.item_checkbox.bind(active=on_checkbox_active)

                # Define o evento de adicionar/remover a label 'Não Conforme' ao lado do checkbox
                def on_checkbox_active2(checkbox, value, layout=item_layout):
                    if value:
                        layout.add_widget(Label(text='Não Conforme', size_hint_x=1,color=(1, 0, 0.2, 1)))
                    else:
                        for widget in layout.children:
                            if isinstance(widget, Label) and widget.text == 'Não Conforme':
                                layout.remove_widget(widget)
                                break

                self.item_checkbox2.bind(active=on_checkbox_active2)

            elif cell_value == 'dx' and cell_value2 == int(self.tipo):
                # Cria um textbox
                item_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
                item_label = Label(text=str(sheet.cell(row=row, column=2).value), halign='right', valign='middle', size_hint_x=None, width='300sp')
                item_textbox = TextInput(multiline=False, size_hint_x=None, width='100sp')
                item_layout.add_widget(item_label)
                item_layout.add_widget(item_textbox)
                components_layout.add_widget(item_layout)

        scroll_view = ScrollView(size_hint=(1, 1))
        scroll_view.add_widget(components_layout)
        self.layout.add_widget(scroll_view)

        # Cria um novo TextInput para observação
        self.observacao_input = TextInput(multiline=False, hint_text='Observações')
        observacao_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height='30sp')
        observacao_layout.add_widget(Label(text='Obs:', size_hint_x=None, width='100sp'))
        observacao_layout.add_widget(self.observacao_input)
        self.layout.add_widget(observacao_layout)

        # Adiciona o botão de exportar PDF após o checklist ser gerado
        export_button = Button(text='Exportar Relatório', size_hint=(None, None), size=(150, 50), pos_hint={'center_x': 0.5})
        export_button.bind(on_press=lambda instance: self.exportar_e_enviar_relatorio3(selected_planilha, selected_coluna))
        self.layout.add_widget(export_button)

    def exportar_e_enviar_relatorio(self, planilha, coluna):
        if not self.verificar_campos_preenchidos():
            # Mostrar mensagem de erro
            popup = Popup(title='Campos não preenchidos',
                          content=Label(text='Por favor, preencha todos os campos antes de exportar.'),
                          size_hint=(None, None), size=(400, 200))
            popup.open()
            return

        # Cria o nome do arquivo PDF com base no tempo
        now = datetime.now()
        dt_string = now.strftime("%Y%m%d")
        file_name2 = f'{self.numero_chassi_input.text}-Relatorio.pdf'  # nome do relatorio chassi
        pdf_path2 = os.path.join(os.getcwd(), file_name2)

        # Verifica se o arquivo PDF já existe
        if os.path.exists(pdf_path2):
            file_name = f'{planilha}_{coluna}-{self.numero_serie_input.text}-Relatorio.pdf'
            pdf_path = os.path.join(os.getcwd(), file_name)
            c = canvas.Canvas(pdf_path)
            self.gerar_pdf(c)
            c.save()
            
        else:
            popup = Popup(title='Relatório não encontrado',
                          content=Label(text=f'O relatório do chassi não existe.'),
                          size_hint=(None, None), size=(400, 200))
            popup.open()
            return

        # Gera o PDF
        #c = canvas.Canvas(pdf_path)
        # self.gerar_pdf(c)

        # Salva o PDF
        #c.save()

        # Envia o e-mail com o PDF anexado
        self.enviar_email_outlook(pdf_path)

        #self.adicionar_ao_pdf_existente(pdf_path, pdf_path2)
        #c.save()
        # Exibir mensagem de arquivo criado
        popup = Popup(title='Relatório Criado',
                      content=Label(text=f'Relatório salvo como {file_name} e enviado por e-mail.'),
                      size_hint=(None, None), size=(400, 200))
        popup.open()

        self.adicionar_ao_pdf_existente(pdf_path, pdf_path2,c)

        # Adicionar dados no Excel
        workbook_path = 'BD.xlsx'  
        workbook = load_workbook(workbook_path)
        sheet = workbook['Chassis']

        # Procura pelo valor selecionado (text) na coluna 1
        next_row = None
        for row in range(2, sheet.max_row + 1):  # Começa de 2 para ignorar o cabeçalho
                if sheet.cell(row=row, column=1).value == str(self.chassi_input.text):
                    next_row = row
                    break
            
        # Se encontrou o valor, atualiza o TextInput nivel com o valor da coluna 2
        # Adicionar dados na próxima linha vazia
        sheet.cell(row=next_row, column=3, value=str(self.numero_serie_input.text))
        sheet.cell(row=next_row, column=4, value=str(self.equipamento_input.text))
        sheet.cell(row=next_row, column=5, value=str(self.modelo_input.text))
        sheet.cell(row=next_row, column=6, value=int('2'))

        # Salvar o workbook
        workbook.save(workbook_path)
        workbook.close()

        self.layout.clear_widgets()  # Limpa todos os widgets atuais
        self.carregar_layout_inicial()  # Reconstrói a interface


        self.layout.clear_widgets()  # Limpa todos os widgets atuais
        self.carregar_layout_inicial()  # Reconstrói a interface

    def adicionar_ao_pdf_existente(self, pdf_path, pdf_path2,c):
        
            # Ler o PDF existente
            reader_existing = PdfReader(pdf_path2)
            reader_new = PdfReader(pdf_path)
        
            # Criar um escritor de PDF
            writer = PdfWriter()

            # Adicionar páginas do PDF existente
            for page_num in range(len(reader_existing.pages)):
                writer.add_page(reader_existing.pages[page_num])

            # Adicionar páginas do novo PDF
            for page_num in range(len(reader_new.pages)):
                writer.add_page(reader_new.pages[page_num])

            # Escrever o novo PDF combinado
            with open(pdf_path, 'wb') as output_pdf:
                writer.write(output_pdf)
            
            os.remove(pdf_path2)    

    def verificar_campos_preenchidos(self):
        if not self.usuario_input.text or not self.numero_serie_input.text:
            return False

        for widget in self.layout.children:
            if isinstance(widget, ScrollView):
                components_layout = widget.children[0]
                for item_layout in components_layout.children:
                    if isinstance(item_layout, BoxLayout):
                        checkboxes = []
                        for child in item_layout.children:
                             # Adiciona CheckBox ao grupo
                            if isinstance(child, CheckBox):
                                    checkboxes.append(child)

                            # Verifica o par de CheckBox
                            if len(checkboxes) == 2:
                               if not checkboxes[0].active and not checkboxes[1].active:
                                  return False
                               
                            if isinstance(child, TextInput) and not child.text:
                                return False
                 
        return True

    def gerar_pdf(self, c):
        # Adicionar cabeçalho
        header_text = self.item_spinner5.text
        c.setFont("Helvetica-Bold", 16)
        c.drawString(100, 800, f'Equipamento: {header_text}')

        # Linha com texto "Relatório de Inspeção"
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredString(300, 740, "Relatório de Inspeção Mecanica")

        # Adicionar uma linha
        c.line(30, 730, 550, 730)

        # Usuário e Número de Série no PDF
        c.setFont("Helvetica", 12)
        c.drawString(100, 780, f'Inspetor: {self.usuario_input.text}')
        c.drawString(100, 760, f'Número de Série: {self.numero_serie_input.text}')

        # Obter todas as labels dos checkboxes marcados e labels dos textboxes com valores
        self.nok=None
        labels_checked = []
        for widget in self.layout.children:
            if isinstance(widget, ScrollView):
                components_layout = widget.children[0]
                for item_layout in components_layout.children:
                    if isinstance(item_layout, BoxLayout):
                        for child in item_layout.children:
                            if isinstance(child, CheckBox)and child.active:
                                #if isinstance(sub_child, Label)and self.item_checkbox2.active==True:
                                     #  labels_checked.append("Não Conforme")
                                     #  labels_checked.append(sub_child.text) 
                                # Adicionar a label associada ao checkbox marcado
                                for sub_child in item_layout.children:
                                    if isinstance(sub_child, Label) and self.item_checkbox.active:
                                       labels_checked.append(sub_child.text)
                                    if isinstance(sub_child, Label) and self.item_checkbox2.active:
                                        labels_checked.append(sub_child.text)
                                            
                            elif isinstance(child, TextInput):
                                # Adicionar a label e o valor do textbox
                                for sub_child in item_layout.children:
                                    if isinstance(sub_child, Label):
                                        labels_checked.append(f"{sub_child.text}: {child.text}")
                                                                    
                
        # Inverter a lista de labels e valores
        if "Não Conforme" in labels_checked:
            self.nok=1
        labels_checked.reverse()

        # Escrever as labels e valores no PDF
        c.setFont("Helvetica", 12)
        y_position = 700
        page_height = 750  # Altura da página para controle de quebra de página
        for label in labels_checked:
            if y_position < 50:  # Se o espaço na página atual acabou, cria uma nova página
                c.showPage()
                c.setFont("Helvetica", 12)
                page_height = 800 # Reinicia a altura da nova página
                y_position = 800

            c.drawString(100, y_position, label)
            y_position -= 20  # Espaçamento entre as labels
            page_height -= 20  # Atualiza a altura da página atual
        
        #inserir observacao
        if y_position < 50:  # Se o espaço na página atual acabou, cria uma nova página
                c.showPage()
                c.setFont("Helvetica", 12)
                page_height = 800 # Reinicia a altura da nova página
                y_position = 800
              
        c.setFont("Helvetica", 12)
        c.drawString(100, y_position, f'Obs: {self.observacao_input.text}')
 
    def exportar_e_enviar_relatorio3(self, planilha, coluna):
        if not self.verificar_campos_preenchidos3():
            # Mostrar mensagem de erro
            popup = Popup(title='Campos não preenchidos',
                          content=Label(text='Por favor, preencha todos os campos antes de exportar.'),
                          size_hint=(None, None), size=(400, 200))
            popup.open()
            return

        # Cria o nome do arquivo PDF com base no tempo
        now = datetime.now()
        dt_string = now.strftime("%Y%m%d")
        file_name2 = f'{planilha}_{coluna}-{self.numero_serie_input.text}-Relatorio.pdf'  # nome do relatorio chassi
        pdf_path2 = os.path.join(os.getcwd(), file_name2)

        # Verifica se o arquivo PDF já existe
        if os.path.exists(pdf_path2):
            file_name = f'{planilha}_{coluna}-{self.numero_serie_input.text}-Relatorio-1.pdf'
            pdf_path = os.path.join(os.getcwd(), file_name)
            c = canvas.Canvas(pdf_path)
            self.gerar_pdf3(c)
            c.save()
            
        else:
            popup = Popup(title='Relatório não encontrado',
                          content=Label(text=f'O relatório de inspeção não existe.'),
                          size_hint=(None, None), size=(400, 200))
            popup.open()
            return

        # Gera o PDF
        #c = canvas.Canvas(pdf_path)
        # self.gerar_pdf(c)

        # Salva o PDF
        #c.save()

        # Envia o e-mail com o PDF anexado
        self.enviar_email_outlook(pdf_path)

        #self.adicionar_ao_pdf_existente(pdf_path, pdf_path2)
        #c.save()
        # Exibir mensagem de arquivo criado
        popup = Popup(title='Relatório Criado',
                      content=Label(text=f'Relatório salvo como {file_name} e enviado por e-mail.'),
                      size_hint=(None, None), size=(400, 200))
        popup.open()

        self.adicionar_ao_pdf_existente3(pdf_path, pdf_path2)

         # Adicionar dados no Excel
        workbook_path = 'BD.xlsx'  
        workbook = load_workbook(workbook_path)
        sheet = workbook['Chassis']

        # Procura pelo valor selecionado (text) na coluna 1
        next_row = None
        for row in range(2, sheet.max_row + 1):  # Começa de 2 para ignorar o cabeçalho
                if sheet.cell(row=row, column=3).value == self.numero_serie_input.text:
                    next_row = row
                    break
            
        # Se encontrou o valor, atualiza o TextInput nivel com o valor da coluna 2
        # Adicionar dados na próxima linha vazia
        sheet.cell(row=next_row, column=3, value=str(self.numero_serie_input.text))
        sheet.cell(row=next_row, column=4, value=str(self.equipamento_input.text))
        sheet.cell(row=next_row, column=5, value=str(self.modelo_input.text))
        sheet.cell(row=next_row, column=6, value=int(self.tipo))

        # Salvar o workbook
        workbook.save(workbook_path)
        workbook.close()



        self.layout.clear_widgets()  # Limpa todos os widgets atuais
        self.carregar_layout_inicial()  # Reconstrói a interface

    def verificar_campos_preenchidos3(self):
        
        for widget in self.layout.children:
            if isinstance(widget, ScrollView):
                components_layout = widget.children[0]
                for item_layout in components_layout.children:
                    if isinstance(item_layout, BoxLayout):
                        checkboxes = []
                        for child in item_layout.children:
                            # Adiciona CheckBox ao grupo
                            if isinstance(child, CheckBox):
                                    checkboxes.append(child)

                            # Verifica o par de CheckBox
                            if len(checkboxes) == 2:
                               if not checkboxes[0].active and not checkboxes[1].active:
                                  return False
                            if isinstance(child, TextInput) and not child.text:
                                return False
                

        return True

    def gerar_pdf3(self, c):
        # Adicionar cabeçalho
        header_text = self.equipamento_input.text
        c.setFont("Helvetica-Bold", 16)
        c.drawString(100, 800, f'Equipamento: {header_text}')
        
        # Linha com texto "Relatório de Inspeção"
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredString(300, 740, f'Relatório de Inspeção {self.nivel.text}')
        
        # Adicionar uma linha
        c.line(30, 730, 550, 730)

        # Usuário e Número de Série no PDF
        c.setFont("Helvetica", 12)
        c.drawString(100, 780, f'Inspetor: {self.usuario_input.text}')
        c.drawString(100, 760, f'Número de Série: {self.numero_serie_input.text}')

        # Obter todas as labels dos checkboxes marcados e labels dos textboxes com valores
        self.nok=None
        labels_checked = []
        for widget in self.layout.children:
            if isinstance(widget, ScrollView):
                components_layout = widget.children[0]
                for item_layout in components_layout.children:
                    if isinstance(item_layout, BoxLayout):
                        for child in item_layout.children:
                            if isinstance(child, CheckBox)and child.active:
                                #if isinstance(sub_child, Label)and self.item_checkbox2.active==True:
                                     #  labels_checked.append("Não Conforme")
                                     #  labels_checked.append(sub_child.text) 
                                # Adicionar a label associada ao checkbox marcado
                                for sub_child in item_layout.children:
                                    if isinstance(sub_child, Label) and self.item_checkbox.active:
                                       labels_checked.append(sub_child.text)
                                    if isinstance(sub_child, Label) and self.item_checkbox2.active:
                                        labels_checked.append(sub_child.text)
                                        #labels_checked.append("Não Conforme")
                                        #labels_checked.append(sub_child.text)   
                                         
                                            
                            elif isinstance(child, TextInput):
                                # Adicionar a label e o valor do textbox
                                for sub_child in item_layout.children:
                                    if isinstance(sub_child, Label):
                                        labels_checked.append(f"{sub_child.text}: {child.text}")
                                                                    
                
        # Inverter a lista de labels e valores
        if "Não Conforme" in labels_checked:
            self.nok=1
        labels_checked.reverse()

        # Escrever as labels e valores no PDF
        c.setFont("Helvetica", 12)
        y_position = 700
        page_height = 750  # Altura da página para controle de quebra de página
        for label in labels_checked:
            if y_position < 50:  # Se o espaço na página atual acabou, cria uma nova página
                c.showPage()
                c.setFont("Helvetica", 12)
                page_height = 800 # Reinicia a altura da nova página
                y_position = 800

            c.drawString(100, y_position, label)
            y_position -= 20  # Espaçamento entre as labels
            page_height -= 20  # Atualiza a altura da página atual
        
        #inserir observacao
        if y_position < 50:  # Se o espaço na página atual acabou, cria uma nova página
                c.showPage()
                c.setFont("Helvetica", 12)
                page_height = 800 # Reinicia a altura da nova página
                y_position = 800
              
        c.setFont("Helvetica", 12)
        c.drawString(100, y_position, f'Obs: {self.observacao_input.text}')

    def adicionar_ao_pdf_existente3(self, pdf_path, pdf_path2):
        
            # Ler o PDF existente
            reader_existing = PdfReader(pdf_path2)
            reader_new = PdfReader(pdf_path)
        
            # Criar um escritor de PDF
            writer = PdfWriter()

            # Adicionar páginas do PDF existente
            for page_num in range(len(reader_existing.pages)):
                writer.add_page(reader_existing.pages[page_num])

            # Adicionar páginas do novo PDF
            for page_num in range(len(reader_new.pages)):
                writer.add_page(reader_new.pages[page_num])

            # Escrever o novo PDF combinado
            with open(pdf_path, 'wb') as output_pdf:
                writer.write(output_pdf)

             # Escrever o novo PDF combinado em um arquivo temporário
            temp_pdf_path = 'temp_combined.pdf'
            with open(temp_pdf_path, 'wb') as output_pdf:
                writer.write(output_pdf)

            # Renomear o arquivo temporário para o nome do pdf_path2
            os.replace(temp_pdf_path, pdf_path2)
    
            # Excluir o arquivo pdf_path original
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
            
            
            #os.remove(pdf_path2)
            #os.rename(pdf_path2, pdf_path)

    def enviar_email_outlook(self, attachment_path):
        try:
            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)
            mail.To = 'processos01@equimatec.ind.br'
            mail.Subject = 'Relatório de Inspeção'
            if self.nok==1:
                mail.Body = 'Segue em anexo o relatório de inspeção gerado. Atenção! Possui não conformidades.'    
            else:
                mail.Body = 'Segue em anexo o relatório de inspeção gerado, sem resalvas'
            mail.Attachments.Add(attachment_path)
            mail.Send()
        except Exception as e:
            # Manipular erros de envio de e-mail aqui
            print(f"Erro ao enviar e-mail: {str(e)}")

if __name__ == '__main__':
    Relatorios().run()
